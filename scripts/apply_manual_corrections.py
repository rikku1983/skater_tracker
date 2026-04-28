"""
Apply manual corrections from data/flagged_times_llm_corrected.xlsx to the DB.

The user colored corrected cells in red (font color FFFF0000).
Corrections fall into these categories:
  - time_text:  re-parse and update Result.time_seconds; clear flag
  - distance_m: update Race.distance_m (shared across all results in the race)
  - division:   update Race.division
  - round:      update Race.round
  - skater_name + others: reassign Result to correct skater/race
  - long_track column set: delete the Result (spurious long track data)

Usage:
    python scripts/apply_manual_corrections.py [--dry-run]
"""
from __future__ import annotations

import argparse
import csv
import logging
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parents[1]))

import openpyxl

from src.db.session import get_session, init_db
from src.db.models import Event, Race, Result, Skater
from src.utils.times import parse_time

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

RED_RGB = "FFFF0000"
XLSX_PATH = Path("data/flagged_times_llm_corrected.xlsx")
CSV_PATH  = Path("data/flagged_times.csv")


def _load_corrections():
    """Return list of (orig_row_dict, red_fields_dict) for every corrected row."""
    with open(CSV_PATH) as f:
        orig_rows = list(csv.DictReader(f))

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    corrections = []
    for row in ws.iter_rows(min_row=2):
        row_idx = row[0].row
        orig = orig_rows[row_idx - 2]      # xlsx row 2 → csv index 0
        reds = {}
        for cell in row:
            if (cell.font and cell.font.color
                    and cell.font.color.type == "rgb"
                    and cell.font.color.rgb == RED_RGB):
                col = headers[cell.column - 1]
                reds[col] = cell.value     # corrected value
        if reds:
            corrections.append((orig, reds))
    return corrections


def _find_result(db, event_id, skater_name, distance_m, division, round_):
    """Locate the Result in the DB using original identifiers.
    Falls back progressively if strict match returns nothing."""
    base = (
        db.query(Result)
        .join(Race, Result.race_id == Race.id)
        .join(Event, Race.event_id == Event.id)
        .join(Skater, Result.skater_id == Skater.id)
        .filter(Event.id == int(event_id), Skater.full_name == skater_name,
                Race.distance_m == int(distance_m))
    )
    for div_filter, round_filter in [
        (division, round_),   # strict
        (division, None),     # relax round
        (None, None),         # relax both
    ]:
        q = base
        if div_filter:
            q = q.filter(Race.division == div_filter)
        if round_filter:
            q = q.filter(Race.round == round_filter)
        results = q.all()
        if len(results) == 1:
            return results[0]
        if len(results) > 1:
            logger.debug("  Ambiguous (%d) for %s event=%s %sm — using first",
                         len(results), skater_name, event_id, distance_m)
            return results[0]
    return None


def _find_or_create_race(db, event_id, division, distance_m, round_, dry_run):
    """Return an existing Race matching the given fields, or create one."""
    race = (
        db.query(Race)
        .filter(
            Race.event_id == int(event_id),
            Race.division == division,
            Race.distance_m == int(distance_m),
            Race.round == round_,
        )
        .first()
    )
    if race:
        return race
    if dry_run:
        logger.info("  [dry-run] Would create Race event=%s %sm %s %s",
                    event_id, distance_m, division, round_)
        return None
    race = Race(event_id=int(event_id), division=division,
                distance_m=int(distance_m), round=round_)
    db.add(race)
    db.flush()
    return race


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    init_db()
    db = get_session()

    corrections = _load_corrections()
    logger.info("Loaded %d corrected rows from Excel", len(corrections))

    stats = {"time": 0, "race": 0, "delete": 0, "reassign": 0, "skip": 0}
    # Track Race objects already updated this session to avoid double-updates
    updated_races: dict[int, dict] = {}

    for orig, reds in corrections:
        event_id   = orig["event_id"]
        skater_name = orig["skater_name"]
        orig_dist  = orig["distance_m"]
        orig_div   = orig["division"]
        orig_round = orig["round"]

        result = _find_result(db, event_id, skater_name, orig_dist, orig_div, orig_round)
        if result is None:
            logger.warning("  NOT FOUND: event=%s %s %sm %s %s",
                           event_id, skater_name, orig_dist, orig_div, orig_round)
            stats["skip"] += 1
            continue

        # --- long track: delete the Result ---
        if reds.get("long_track") or reds.get("long track"):
            logger.info("  DELETE (long track): event=%s %s %sm", event_id, skater_name, orig_dist)
            if not args.dry_run:
                db.delete(result)
            stats["delete"] += 1
            continue

        # --- skater reassignment (+ race fields) ---
        if "skater_name" in reds:
            new_name  = reds["skater_name"]
            new_dist  = int(reds.get("distance_m", orig_dist))
            new_div   = reds.get("division", orig_div)
            new_round = reds.get("round", orig_round)

            new_skater = (
                db.query(Skater)
                .filter(Skater.full_name == new_name)
                .first()
            )
            if new_skater is None:
                logger.warning("  Skater not found: %s — skipping reassignment", new_name)
                stats["skip"] += 1
                continue

            new_race = _find_or_create_race(db, event_id, new_div, new_dist, new_round, args.dry_run)
            logger.info("  REASSIGN: event=%s %s -> %s | race %sm %s",
                        event_id, skater_name, new_name, new_dist, new_div)
            if not args.dry_run and new_race:
                result.skater_id = new_skater.id
                result.race_id   = new_race.id
                result.data_flags = None
            stats["reassign"] += 1
            continue

        # --- Race-level field corrections (distance_m / division / round) ---
        # Move the Result to the correct target Race rather than mutating the
        # source Race (which might still be correct for other results in it).
        race_changed = any(f in reds for f in ("distance_m", "division", "round"))
        if race_changed:
            src_race = result.race
            new_dist  = int(reds["distance_m"]) if "distance_m" in reds else src_race.distance_m
            new_div   = reds.get("division",  src_race.division)
            new_round = reds.get("round",     src_race.round)

            target_race = _find_or_create_race(
                db, event_id, new_div, new_dist, new_round, args.dry_run)
            logger.info("  MOVE result: event=%s %s | race %sm '%s' -> %sm '%s'",
                        event_id, skater_name,
                        src_race.distance_m, src_race.division,
                        new_dist, new_div)
            if not args.dry_run and target_race:
                result.race_id = target_race.id
            stats["race"] += 1

        # --- Result time correction ---
        if "time_text" in reds:
            new_time_text = str(reds["time_text"])
            new_time_s    = parse_time(new_time_text)
            logger.info("  TIME: event=%s %s %s -> %s (%.3fs)",
                        event_id, skater_name, orig["time_text"], new_time_text, new_time_s or 0)
            if not args.dry_run:
                result.time_text    = new_time_text
                result.time_seconds = new_time_s
            stats["time"] += 1

        # Clear the flag if any correction was applied
        if not args.dry_run:
            result.data_flags = None

    if not args.dry_run:
        db.commit()
        logger.info("Committed all changes.")

    logger.info("")
    logger.info("=== Summary ===")
    for k, v in stats.items():
        logger.info("  %-12s %d", k, v)
    db.close()


if __name__ == "__main__":
    main()
